"""
Chapter Four - Step 1: Merge the two uploaded Excel feature batches into a single CSV.

The Elliptic Bitcoin Dataset features file was too large to upload as one file, so it
was split by the researcher into two Excel batches before upload. This script reads
both batches back out, in the original row order, and writes them to a single CSV
for downstream processing.
"""
import os
os.makedirs("outputs", exist_ok=True)
import csv
import time
import openpyxl

BATCH_FILES = [
    "data/elliptic_txs_features_batch1.xlsx",
    "data/elliptic_txs_features_batch2.xlsx",
]
OUTPUT_CSV = "outputs/ch4_features_merged.csv"

t0 = time.time()
out_file = open(OUTPUT_CSV, "w", newline="")
writer = csv.writer(out_file)

total_rows = 0
for batch_path in BATCH_FILES:
    workbook = openpyxl.load_workbook(batch_path, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows_in_batch = 0
    for row in sheet.iter_rows(values_only=True):
        writer.writerow(row)
        rows_in_batch += 1
    workbook.close()
    total_rows += rows_in_batch
    print(f"{batch_path}: {rows_in_batch} rows merged (elapsed {time.time() - t0:.1f}s)")

out_file.close()
print(f"TOTAL ROWS WRITTEN: {total_rows}")
print(f"Merged CSV saved to: {OUTPUT_CSV}")
print(f"Done in {time.time() - t0:.1f}s")
